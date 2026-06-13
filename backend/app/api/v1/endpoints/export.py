"""
Export endpoint — CSV and XLSX download for a conversation turn's query result.

Spec: EX-001, EX-002, EX-003
  - EX-001: GET /conversations/{conv_id}/turns/{turn_id}/export?format=csv|xlsx
  - EX-002: Rows fetched from the turn record stored in the DB (query_result JSON column)
  - EX-003: Filename = "{tenant_slug}_{turn_id_short}.{format}"
  - EX-004: Max 10,000 rows exported; truncation header X-Export-Truncated: true
  - EX-005: XLSX uses openpyxl; CSV uses stdlib csv (no extra deps for CSV)
  - EX-006: Row data is sanitised — formula injection prevention (cells starting with
            = + - @ are prefixed with a single quote in CSV, space in XLSX cell value)

Security: auth via get_current_user, RLS via tenant_id filter on turn lookup.
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_user
from app.core.logging import get_logger
from app.db.session import get_db
from app.models.conversation import Conversation, ConversationTurn

log = get_logger("export")
router = APIRouter(tags=["export"])

_MAX_ROWS = 10_000
_FORMULA_CHARS = {"=", "+", "-", "@"}


def _sanitise_cell(value: Any) -> str:
    """Convert value to string, neutralise formula-injection prefixes."""
    s = "" if value is None else str(value)
    if s and s[0] in _FORMULA_CHARS:
        s = "'" + s  # prefix with single quote for CSV
    return s


def _sanitise_xlsx_cell(value: Any) -> Any:
    """For XLSX, prefix formula-injection strings with a space."""
    if value is None:
        return ""
    s = str(value)
    if s and s[0] in _FORMULA_CHARS:
        return " " + s
    return s


@router.get("/conversations/{conversation_id}/turns/{turn_id}/export")
async def export_turn_data(
    conversation_id: uuid.UUID,
    turn_id: uuid.UUID,
    format: str = Query(default="csv", pattern="^(csv|xlsx)$"),
    current_user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tenant_id = current_user.tenant_id

    # Load turn — RLS via tenant_id join through conversation
    result = await db.execute(
        select(ConversationTurn)
        .join(Conversation, Conversation.id == ConversationTurn.conversation_id)
        .where(
            ConversationTurn.id == turn_id,
            ConversationTurn.conversation_id == conversation_id,
            Conversation.tenant_id == tenant_id,
        )
    )
    turn = result.scalar_one_or_none()
    if turn is None:
        raise HTTPException(status_code=404, detail="Turn not found")

    raw = turn.answer_data or {}
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except json.JSONDecodeError:
            raw = {}

    rows: list[dict] = raw.get("rows", [])
    columns: list[str] = raw.get("columns", [])

    if not rows or not columns:
        raise HTTPException(status_code=422, detail="No tabular data available for this turn")

    truncated = len(rows) > _MAX_ROWS
    rows = rows[:_MAX_ROWS]

    turn_id_short = str(turn_id).split("-")[0]

    headers_extra: dict[str, str] = {}
    if truncated:
        headers_extra["X-Export-Truncated"] = "true"
        headers_extra["X-Export-Row-Limit"] = str(_MAX_ROWS)

    if format == "csv":
        return _build_csv(rows, columns, turn_id_short, headers_extra)
    return _build_xlsx(rows, columns, turn_id_short, headers_extra)


def _build_csv(
    rows: list[dict],
    columns: list[str],
    name_suffix: str,
    extra_headers: dict,
) -> StreamingResponse:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([_sanitise_cell(row.get(c)) for c in columns])
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="export_{name_suffix}.csv"',
            **extra_headers,
        },
    )


def _build_xlsx(
    rows: list[dict],
    columns: list[str],
    name_suffix: str,
    extra_headers: dict,
) -> StreamingResponse:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="XLSX export requires openpyxl. Install it with: pip install openpyxl",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"

    # Header row — bold, light blue background
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEEFF")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(
                row=row_idx,
                column=col_idx,
                value=_sanitise_xlsx_cell(row.get(col_name)),
            )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="export_{name_suffix}.xlsx"',
            **extra_headers,
        },
    )
